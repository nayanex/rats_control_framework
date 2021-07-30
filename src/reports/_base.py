from os import path

from openpyxl import Workbook, load_workbook

from src.automation.adapters.mixins import ValuesMixin
from src.config import get_output_filepath, get_template_filepath
from src.utils.logger import CustomLogger
from src.utils.utils import get_month_year_str, get_quarter_year_str

logger = CustomLogger(__name__).get_logger()


class BaseReport:
    FILENAME = None
    report_type = None
    evidences = None
    report_queries = None
    general_data = None

    class ReportType(ValuesMixin):
        MONTHLY = "monthly"
        QUARTERLY = "quarterly"

    def __init__(self, month, year, db):
        self.month = month
        self.year = year
        self.db = db
        self.wb = self.get_template_file()
        self.check_requirements_initialization()

    def check_requirements_initialization(self):
        """
        Check required values are properly initialized.
        """
        if not self.report_queries:
            raise AttributeError("Please specify report_queries class")

        if self.report_type not in (self.ReportType.values()):
            raise ValueError("Invalid report_type value")

    def get_template_file(self) -> Workbook:
        """
        Get template file for chosen report
        :return: worksheet object
        """
        filepath = get_template_filepath(self.FILENAME)

        if path.isfile(filepath):
            return load_workbook(filepath)

        return Workbook()

    def run_evidences_generator(self, wb):
        """
        Generate report evidences.
        """
        for evidence in self.evidences:
            logger.info("************ {} ************".format(evidence.__name__))
            evidence = evidence(
                self.month,
                self.year,
                self.general_data,
                wb,
                self.report_queries(self.db),
            )
            logger.info("Get values")
            evidence.get_values()
            logger.info("Update worksheet with values")
            evidence.update_worksheet_with_values()

    def populate_evidences_pages(self):
        """
        Populate all evidence pages
        """
        logger.info("Start automation process")
        logger.info("############ {} ############".format(self.FILENAME))
        logger.info("Time range: {}-{}".format(self.month, self.year))
        self.run_evidences_generator(self.wb)
        logger.info("Saving output")
        self.save_output()
        logger.info("#" * 64)
        logger.info("Report generation successfully done")

    def get_time_range_str(self):
        """
        Provide a string with processed time range depends on report type
        Report type can be `month` or `quarter`
        """
        if self.report_type == self.ReportType.MONTHLY:
            return get_month_year_str(self.month, self.year)
        elif self.report_type == self.ReportType.QUARTERLY:
            return get_quarter_year_str(self.month, self.year)

    def save_output(self):
        """
        Save new report in output folder with name containing time range
        """
        report_name = self.FILENAME
        time_range = self.get_time_range_str()
        self.wb.save(get_output_filepath(report_name, time_range))

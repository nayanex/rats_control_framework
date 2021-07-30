from abc import abstractmethod

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Border, Font, Side

from src.automation.adapters.delivery_system_map import DELIVERY_SYSTEM_MAP
from src.automation.data_access_layer.base_queries import Query
from src.automation.data_transfer_object.rats_controls import (
    GeneralData,
    QueryResult,
    RequiredData,
)
from src.utils.logger import CustomLogger
from src.utils.utils import (
    create_dict_list_from_tuple_props,
    get_quarter_from_month,
    merge_list_of_dictionaries,
    remap_dict_list,
)

logger = CustomLogger(__name__).get_logger()


class BaseEvidence:
    """Base class to generate evidences for Control Framework reports."""

    worksheet_name = "base"
    ws = None
    values = None
    general_data = None
    month = None
    year = None

    def __init__(
        self,
        month: int,
        year: int,
        general_data: RequiredData,
        wb: Workbook,
        query: Query,
    ):
        self.ws = self.get_worksheet(wb)
        self.month = month
        self.year = year
        self.query = query
        self.general_data = general_data
        self.quarter = get_quarter_from_month(self.month)
        self.fetch_required_data()

    def get_worksheet(self, wb):
        if self.worksheet_name in wb:
            return wb[self.worksheet_name]
        ws = wb.active
        ws.title = self.worksheet_name
        return ws

    @abstractmethod
    def get_values(self):
        raise NotImplementedError

    @abstractmethod
    def update_worksheet_with_values(self):
        raise NotImplementedError

    def fetch_required_data(self) -> GeneralData:
        """
        General data used between reports.
        """
        if not hasattr(self.general_data, "calculation_requests"):
            request_ids = self.query.get_calc_request_ids(self.month, self.year).data

            if not request_ids:
                msg_error = "There are no calculation requests available for the date: {}-{}.".format(
                    self.year, self.month
                )
                logger.exception(msg_error)
                raise ValueError(msg_error)

            self.general_data.calculation_requests = self.query.get_calc_requests(
                request_ids
            )

            src_x_req = create_dict_list_from_tuple_props(
                self.general_data.calculation_requests.data,
                "delivering_system",
                "calculation_request_id",
            )
            src_x_req = remap_dict_list(src_x_req, DELIVERY_SYSTEM_MAP)
            self.general_data.source_per_request = merge_list_of_dictionaries(src_x_req)

        return self.general_data

    def add_query_and_result(
        self, start_cell: Cell, result: QueryResult, dict_params: dict
    ):
        """
        Add SQL query and its result to worksheet
        :param start_cell: starting point to print the data
        :param result: dict with query and data inside
        :param dict_params: parameters of data to be printed. see `add_dictionary_to_worksheet` method
        """
        column_number = start_cell.column
        row_number = start_cell.row
        bold_font = Font(bold=True)
        cell = self.ws.cell(row=row_number, column=column_number, value="Query:")
        cell.font = bold_font
        self.ws.cell(row=row_number + 1, column=column_number, value=result.query_cmd)
        cell = self.ws.cell(row=row_number + 2, column=column_number, value="Result:")
        self.ws.cell(row=row_number + 2, column=column_number, value="Result:")
        cell.font = bold_font
        self.add_dictionary_to_worksheet(
            self.ws.cell(row=row_number + 3, column=column_number),
            result.data,
            **dict_params
        )

    def add_dictionary_to_worksheet(
        self, start_cell: Cell, data: list, bold_columns=None
    ):
        """
        Add list of dictionaries to worksheet from the starting point.
        If value of dict is tuple - first element is value, second is color of text
        :param start_cell: cell object from which table will be printed
        :param data: dict
        :param bold_columns: list of columns to be printed bold
        """
        if not bold_columns:
            bold_columns = list()
        headers = list(data[0])
        bold_column_indices = {headers.index(i) for i in bold_columns}
        column_number = start_cell.column
        row_number = start_cell.row
        side = Side(style="thin")

        def _print_row(row_values):
            cell = self.ws.cell(row=row_number, column=column_number + i)
            if type(row_values[i]) == tuple:
                cell.value = row_values[i][0]
                color = row_values[i][1]
            else:
                cell.value = row_values[i]
                color = None
            cell.font = Font(color=color, bold=i in bold_column_indices)

            cell.border = Border(
                top=side if row_number == start_row else None,
                left=side if i == 0 else None,
                right=side if i == len(row_values) - 1 else None,
                bottom=side if row_number == last_row else None,
            )

        start_row = start_cell.row
        last_row = start_row + 1

        for i in range(len(headers)):
            _print_row(headers)

        start_row = row_number + 1
        last_row = row_number + len(data)

        for row in data:
            row_number += 1
            values = list(row.values())
            for i in range(len(values)):
                _print_row(values)

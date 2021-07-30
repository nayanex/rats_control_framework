from openpyxl.styles import Font

from src.config import get_latest_mev_filename
from src.evidences._base import BaseEvidence
from src.utils.utils import get_scenario_name


class EvidenceC(BaseEvidence):
    worksheet_name = "evidence C"

    CODE_NL_GDP_CONA = "NL_GDP_CONA"
    CODE_NL_PRIVATE_CONSUMTION_CONA = "NL_PRIVATE_CONSUMPTION_CONA"
    CODE_NL_PRIVATE_INVESTMENT_CONA = "NL_PRIVATE_INVESTMENT_CONA"

    code_properties = {
        CODE_NL_GDP_CONA: {
            "column_number": 5,
            "color_code": "cc00cc",
        },
        CODE_NL_PRIVATE_CONSUMTION_CONA: {
            "column_number": 6,
            "color_code": "00cc00",
        },
        CODE_NL_PRIVATE_INVESTMENT_CONA: {
            "column_number": 7,
            "color_code": "ff0000",
        },
    }

    def get_values(self):
        """
        Get all required data from DB
        """
        self.values = self.general_data
        self.values.mps_mevs = self.query.get_mevs(
            calc_id=self.general_data.mdm_dataset.data[0]["id"],
            quarter=self.quarter,
            year=self.year,
            codes=list(self.code_properties.keys()),
        )
        self.values.economic_scenario_projections = (
            self.get_economic_scenario_projections()
        )

    def update_worksheet_with_values(self):
        """
        Provide received data to worksheet as evidences
        """
        self._print_mevs_data()
        self._print_economic_scenario_projections()

    def get_economic_scenario_projections(self):
        filepath = get_latest_mev_filename()
        compare_attributes = self.compare_attributes()
        line_number = 0
        result = []
        with open(file=filepath) as f:
            for line in f:
                line_number += 1
                attributes = line.split("|")
                if all([attributes[i] == v for i, v in compare_attributes.items()]):
                    result.append((None, ["..............."]))
                    result.append((line_number, attributes))
                    result.append((None, ["..............."]))
                    break
                elif line_number <= 4:
                    result.append((line_number, attributes))
            else:
                raise ValueError(
                    "The line in economic_scenario_projections with required attributes has not found. Attributes: {}".format(
                        compare_attributes.values()
                    )
                )
            return result

    def compare_attributes(self):
        """
        Compare data in MEV file with the result of db query.
        Try to find full match of attributes.
        Return few first rows and full match row.
        """
        result = {
            0: get_scenario_name(self.quarter, self.year),
            3: str(self.quarter),
            4: "projected",
        }
        for prop in self.code_properties:
            for line in self.values.mps_mevs.data:
                if prop == line["code"]:
                    result[self.code_properties[prop]["column_number"]] = str(
                        line["value"]
                    )
        return result

    def _print_mevs_data(self):
        mps_mevs = self.values.mps_mevs
        for row in mps_mevs.data:
            row["value"] = (
                row["value"],
                self.code_properties.get(row["code"])["color_code"],
            )
        self.ws["B09"] = (
            "A query is used to fetch the contents of the MEV table which was imported in CORE and used "
            "in the calculation by RATS."
        )
        self.add_query_and_result(
            self.ws["B12"],
            mps_mevs,
            dict_params={"bold_columns": ["import_dataset_id"]},
        )

    def _print_economic_scenario_projections(self):
        scenarios_data = self.values.economic_scenario_projections
        self.ws[
            "B21"
        ] = "Below data shows the content of MEV table 2078 of {} {}".format(
            self.quarter, self.year
        )
        self.ws["B21"].font = Font(bold=True)
        row_number = 24
        for line_number, line in scenarios_data:
            self.ws.cell(row=row_number, column=1, value=line_number)
            column = 1
            for value in line:
                column += 1
                self.ws.cell(row=row_number, column=column, value=value)
            row_number += 1
        # highlight important cells
        for prop in self.code_properties.values():
            cell = self.ws.cell(row=row_number - 2, column=prop["column_number"] + 2)
            cell.font = Font(color=prop["color_code"])

        cell = self.ws.cell(
            row=row_number + 2,
            column=2,
        )
        cell.value = "Conclusion: Above sample test proves the correct table is used"
        cell.font = Font(bold=True)
